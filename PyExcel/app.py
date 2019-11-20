import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
ws = wb.active
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a1"]
print(cell.value)
print(cell.row)
print(cell.column)
print(cell.coordinate)

cell = sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet["a"]
print(column)

cells = sheet["a:c"]
print(cells)


sheet.append([1, 2, 3])
# sheet.insert_row()
ws["e5"] = ("ciao")
wb.save("transaction2.xlsx")
